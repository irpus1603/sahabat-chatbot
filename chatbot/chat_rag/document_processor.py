"""
Document Processing Module for RAG System

This module handles document processing, including reading different file formats
and extracting text content for embedding generation.
"""

import os
import logging
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from django.conf import settings
from docling.document_converter import DocumentConverter

# Document processing libraries
import PyPDF2
try:
    import docx2txt
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False
    
import openpyxl
from openpyxl import load_workbook

from django.db import models
from .models import RAGDocument

logger = logging.getLogger(__name__)

class DocumentProcessor:
    """Process various document formats and extract text content"""
    
    def __init__(self):
        self.supported_formats = settings.RAG_CONFIG['SUPPORTED_FORMATS']
        self.document_folder = settings.RAG_CONFIG['DOCUMENT_FOLDER']
        
        # Ensure document folder exists
        self.document_folder.mkdir(exist_ok=True)
    
    def process_folder(self, folder_path: Optional[Path] = None, recursive: bool = True) -> List[RAGDocument]:
        """Process all documents in a folder (recursively by default)"""
        if folder_path is None:
            folder_path = self.document_folder
        
        processed_docs = []
        
        if not folder_path.exists():
            logger.warning(f"Document folder {folder_path} does not exist")
            return processed_docs
        
        if recursive:
            # Walk through all subdirectories recursively
            for file_path in folder_path.rglob('*'):
                if file_path.is_file() and file_path.suffix.lower() in self.supported_formats:
                    try:
                        doc = self.process_file(file_path)
                        if doc:
                            processed_docs.append(doc)
                            logger.info(f"Processed: {file_path.relative_to(folder_path)}")
                    except Exception as e:
                        logger.error(f"Error processing {file_path}: {str(e)}")
        else:
            # Only process files in the current directory
            for file_path in folder_path.iterdir():
                if file_path.is_file() and file_path.suffix.lower() in self.supported_formats:
                    try:
                        doc = self.process_file(file_path)
                        if doc:
                            processed_docs.append(doc)
                            logger.info(f"Processed: {file_path.name}")
                    except Exception as e:
                        logger.error(f"Error processing {file_path}: {str(e)}")
        
        logger.info(f"Completed processing folder {folder_path}: {len(processed_docs)} documents processed")
        return processed_docs
    
    def process_file(self, file_path: Path) -> Optional[RAGDocument]:
        """Process a single file and create RAGDocument"""
        try:
            # Check if document already exists
            existing_doc = RAGDocument.objects.filter(
                source=str(file_path),
                is_active=True
            ).first()
            
            if existing_doc:
                logger.info(f"Document {file_path} already exists, skipping")
                return existing_doc
            
            # Extract text based on file type
            content = self._extract_text(file_path)
            #content = self._extract_text_docling(file_path)
            print(f"Extracted content from {file_path}: {content[:100]}...")  # Log first 100 chars
            if not content:
                logger.warning(f"No content extracted from {file_path}")
                return None
            
            # Create RAGDocument
            doc = RAGDocument.objects.create(
                title=file_path.stem,
                content=content,
                source=str(file_path),
                metadata={
                    'file_size': file_path.stat().st_size,
                    'file_type': file_path.suffix.lower(),
                    'original_filename': file_path.name
                }
            )
            
            logger.info(f"Successfully processed {file_path}")
            return doc
            
        except Exception as e:
            logger.error(f"Error processing file {file_path}: {str(e)}")
            return None
    
    def _extract_text(self, file_path: Path) -> str:
        """Extract text content from file based on its type"""
        file_extension = file_path.suffix.lower()
        
        if file_extension == '.pdf':
            return self._extract_pdf_text(file_path)
        elif file_extension == '.docx':
            return self._extract_docx_text(file_path)
        elif file_extension == '.xlsx':
            return self._extract_xlsx_text(file_path)
        elif file_extension in ['.txt', '.md']:
            return self._extract_text_file(file_path)
        else:
            logger.warning(f"Unsupported file type: {file_extension}")
            return ""
    
    def _extract_pdf_text(self, file_path: Path) -> str:
        """Extract text from PDF file"""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = []
                
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    text.append(page.extract_text())
                
                return '\n'.join(text)
        except Exception as e:
            logger.error(f"Error extracting PDF text from {file_path}: {str(e)}")
            return ""
    
    def _extract_docx_text(self, file_path: Path) -> str:
        """Extract text from DOCX file"""
        if not DOCX_AVAILABLE:
            logger.warning("docx2txt not available, skipping DOCX extraction")
            return ""
            
        try:
            text = docx2txt.process(str(file_path))
            return text
        except Exception as e:
            logger.error(f"Error extracting DOCX text from {file_path}: {str(e)}")
            return ""
    
    def _extract_xlsx_text(self, file_path: Path) -> str:
        """Extract text from XLSX file"""
        try:
            workbook = load_workbook(file_path)
            text = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text.append(f"Sheet: {sheet_name}")
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = []
                    for cell in row:
                        if cell is not None:
                            row_text.append(str(cell))
                    if row_text:
                        text.append(' | '.join(row_text))
                
                text.append("")  # Add empty line between sheets
            
            return '\n'.join(text)
        except Exception as e:
            logger.error(f"Error extracting XLSX text from {file_path}: {str(e)}")
            return ""
    
    def _extract_text_file(self, file_path: Path) -> str:
        """Extract text from plain text or markdown file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        except UnicodeDecodeError:
            # Try with different encoding
            try:
                with open(file_path, 'r', encoding='latin-1') as file:
                    return file.read()
            except Exception as e:
                logger.error(f"Error reading text file {file_path}: {str(e)}")
                return ""
        except Exception as e:
            logger.error(f"Error extracting text from {file_path}: {str(e)}")
            return ""
    
    def _extract_text_docling(self, file_path:Path) -> str:
        """Extract text using docling library"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        except UnicodeDecodeError:
            try:
                converter = DocumentConverter()
                result = converter.convert(file_path)
                document = result.document
                
                if result and 'text' in result:
                    return document

            except Exception as e:
                logger.error(f"Error reading docling file {file_path}: {str(e)}")
                return ""
        except Exception as e:
            logger.error(f"Error extracting text from docling file {file_path}: {str(e)}")
            return ""

    def scan_folder(self, folder_path: Optional[Path] = None) -> Dict:
        """Scan folder and return statistics about available documents"""
        if folder_path is None:
            folder_path = self.document_folder
            
        stats = {
            'total_files': 0,
            'supported_files': 0,
            'by_extension': {},
            'folder_structure': [],
            'unprocessed_files': []
        }
        
        if not folder_path.exists():
            return stats
            
        # Get existing processed documents
        processed_sources = set(
            RAGDocument.objects.filter(is_active=True)
            .values_list('source', flat=True)
        )
        
        # Walk through all files recursively
        for file_path in folder_path.rglob('*'):
            if file_path.is_file():
                stats['total_files'] += 1
                extension = file_path.suffix.lower()
                
                # Count by extension
                stats['by_extension'][extension] = stats['by_extension'].get(extension, 0) + 1
                
                # Check if supported
                if extension in self.supported_formats:
                    stats['supported_files'] += 1
                    
                    # Check if already processed
                    if str(file_path) not in processed_sources:
                        relative_path = file_path.relative_to(folder_path)
                        stats['unprocessed_files'].append({
                            'path': str(relative_path),
                            'full_path': str(file_path),
                            'size': file_path.stat().st_size,
                            'size_formatted': self._format_file_size(file_path.stat().st_size),
                            'extension': extension
                        })
        
        # Build folder structure
        dirs = set()
        for file_path in folder_path.rglob('*'):
            if file_path.is_dir():
                relative_path = file_path.relative_to(folder_path)
                if str(relative_path) != '.':
                    dirs.add(str(relative_path))
        
        stats['folder_structure'] = sorted(list(dirs))
        
        return stats
    
    def _format_file_size(self, size_bytes: int) -> str:
        """Format file size in human readable format"""
        if size_bytes == 0:
            return "0 B"
        size_names = ["B", "KB", "MB", "GB", "TB"]
        i = 0
        while size_bytes >= 1024 and i < len(size_names) - 1:
            size_bytes /= 1024.0
            i += 1
        return f"{size_bytes:.1f} {size_names[i]}"
    
    def get_processing_stats(self) -> Dict:
        """Get statistics about document processing"""
        return {
            'total_documents': RAGDocument.objects.filter(is_active=True).count(),
            'by_status': {
                status: count for status, count in 
                RAGDocument.objects.values_list('embedding_status')
                .annotate(count=models.Count('id'))
            },
            'by_file_type': {
                file_type: count for file_type, count in
                RAGDocument.objects.filter(is_active=True)
                .extra(select={'file_type': "JSON_EXTRACT(metadata, '$.file_type')"})
                .values_list('file_type')
                .annotate(count=models.Count('id'))
            }
        }

# Global document processor instance
document_processor = DocumentProcessor()